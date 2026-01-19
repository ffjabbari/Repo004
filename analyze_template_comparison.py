#!/usr/bin/env python3
"""
Compare template Excel file with database to identify missing information.
"""

import openpyxl
import sqlite3
from collections import defaultdict


def analyze_template_structure(template_path):
    """Analyze the template Excel structure."""
    wb = openpyxl.load_workbook(template_path, data_only=True)
    
    template_data = {
        'sheets': {},
        'categories': defaultdict(list)
    }
    
    print("=" * 80)
    print("ANALYZING TEMPLATE STRUCTURE")
    print("=" * 80)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        sheet_info = {
            'name': sheet_name,
            'headers': [],
            'data_rows': [],
            'categories': []
        }
        
        # Get headers (usually row 1)
        headers = []
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header_text = str(cell.value).strip()
                headers.append(header_text)
                if header_text and len(header_text) > 2:
                    sheet_info['categories'].append(header_text)
        
        sheet_info['headers'] = headers
        print(f"  Headers found: {len(headers)}")
        print(f"  Sample headers: {headers[:10]}")
        
        # Look for data rows
        for row_idx in range(2, min(50, sheet.max_row + 1)):
            row = sheet[row_idx]
            row_data = []
            has_data = False
            for cell in row:
                if cell.value is not None:
                    has_data = True
                    row_data.append(cell.value)
            if has_data:
                sheet_info['data_rows'].append(row_data[:10])  # First 10 columns
        
        template_data['sheets'][sheet_name] = sheet_info
    
    wb.close()
    return template_data


def analyze_database_structure(db_path):
    """Analyze what's in the database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    print("\n" + "=" * 80)
    print("ANALYZING DATABASE STRUCTURE")
    print("=" * 80)
    
    db_data = {
        'entities': {},
        'categories': defaultdict(int),
        'income_types': [],
        'expense_types': []
    }
    
    # Get all entities
    cursor.execute("""
        SELECT entity_name, total_amount, transaction_count, category
        FROM EntitySummary
        ORDER BY ABS(total_amount) DESC
    """)
    
    entities = cursor.fetchall()
    print(f"\nTotal entities in database: {len(entities)}")
    
    # Categorize entities
    for entity, amount, count, category in entities:
        db_data['entities'][entity] = {
            'amount': amount,
            'count': count,
            'category': category
        }
        
        # Categorize by prefix
        if entity.startswith('W-2'):
            db_data['income_types'].append('W-2 Income')
        elif entity.startswith('RENTAL-'):
            db_data['income_types'].append('Rental Income')
        elif entity.startswith('PAYPAL-'):
            db_data['expense_types'].append('PayPal Payments')
        elif entity.startswith('UTIL-'):
            db_data['expense_types'].append('Utilities')
        elif entity.startswith('PROPERTY-'):
            if 'TAX' in entity:
                db_data['expense_types'].append('Property Taxes')
            elif 'INSURANCE' in entity:
                db_data['expense_types'].append('Property Insurance')
        elif entity.startswith('STORM-'):
            db_data['expense_types'].append('Storm Expenses')
    
    # Get unique categories
    cursor.execute("SELECT DISTINCT category FROM EntitySummary WHERE category IS NOT NULL")
    categories = [row[0] for row in cursor.fetchall()]
    db_data['categories_list'] = categories
    
    conn.close()
    return db_data


def compare_template_vs_database(template_data, db_data):
    """Compare template requirements with database content."""
    print("\n" + "=" * 80)
    print("COMPARISON: TEMPLATE vs DATABASE")
    print("=" * 80)
    
    missing_info = []
    
    # Check key sheets from template
    template_sheets = template_data['sheets']
    
    print("\n1. INCOME SECTION:")
    print("-" * 80)
    
    # Check for W-2 income
    has_w2 = any('W-2' in entity for entity in db_data['entities'].keys())
    if has_w2:
        print("✓ W-2 Income: Found in database")
    else:
        print("✗ W-2 Income: MISSING")
        missing_info.append("W-2 Income information")
    
    # Check for rental income
    rental_entities = [e for e in db_data['entities'].keys() if e.startswith('RENTAL-')]
    print(f"✓ Rental Income: Found {len(rental_entities)} properties")
    
    # Check for other income (interest, dividends, etc.)
    print("? Other Income (Interest, Dividends, etc.): Need to verify")
    missing_info.append("Other income sources (interest, dividends, capital gains, etc.)")
    
    print("\n2. EXPENSES SECTION:")
    print("-" * 80)
    
    # Check for business expenses (FFJ Consulting LLC)
    if 'FFJ Consulting LLC EXP' in template_sheets:
        print("? Business Expenses (FFJ Consulting LLC): Need to categorize from transactions")
        missing_info.append("Business expense categorization (Software, Hardware, Office, Auto, etc.)")
    
    # Check for property expenses
    property_taxes = [e for e in db_data['entities'].keys() if 'PROPERTY' in e and 'TAX' in e]
    property_insurance = [e for e in db_data['entities'].keys() if 'PROPERTY' in e and 'INSURANCE' in e]
    print(f"✓ Property Taxes: Found {len(property_taxes)} properties")
    print(f"✓ Property Insurance: Found {len(property_insurance)} properties")
    
    # Check for utilities
    util_entities = [e for e in db_data['entities'].keys() if e.startswith('UTIL-')]
    print(f"✓ Utilities: Found {len(util_entities)} utility types")
    
    # Check for rental expenses
    if 'worksheet Rental DetailRecieptA' in template_sheets:
        print("? Rental Property Expenses (Maintenance, Improvements): Need to categorize")
        missing_info.append("Rental property maintenance and improvement expenses by property")
    
    print("\n3. TAX INFORMATION:")
    print("-" * 80)
    
    # Check for tax withholdings
    print("? Tax Withholdings (Federal, State, SSN, Medicare): Need to verify from W-2")
    missing_info.append("Tax withholding details (should be in W-2, but need to verify)")
    
    print("\n4. OTHER TEMPLATE REQUIREMENTS:")
    print("-" * 80)
    
    # Check what other sheets exist in template
    for sheet_name in template_sheets.keys():
        if 'TAX' in sheet_name.upper() or '1099' in sheet_name.upper():
            print(f"? {sheet_name}: Need to verify if we have this information")
            missing_info.append(f"Information for {sheet_name} sheet")
    
    return missing_info


def main():
    template_path = "/Users/fjabbari/@@@PUBLIC/@@@TAX2025/final001.xlsx"
    db_path = "tax_data_2025.db"
    
    # Analyze template
    template_data = analyze_template_structure(template_path)
    
    # Analyze database
    db_data = analyze_database_structure(db_path)
    
    # Compare
    missing_info = compare_template_vs_database(template_data, db_data)
    
    # Summary
    print("\n" + "=" * 80)
    print("SUMMARY: MISSING INFORMATION NEEDED")
    print("=" * 80)
    
    print("\nTo create final002.xlsx for 2025, we need:")
    print("-" * 80)
    
    for i, item in enumerate(missing_info, 1):
        print(f"{i}. {item}")
    
    print("\n" + "=" * 80)
    print("DETAILED ANALYSIS")
    print("=" * 80)
    
    # Show what we have vs what template needs
    print("\nWhat we HAVE in database:")
    print("  ✓ W-2 Income")
    print("  ✓ Rental Income (6 properties)")
    print("  ✓ Property Taxes (7 properties)")
    print("  ✓ Property Insurance (6 properties)")
    print("  ✓ Utilities (6 types)")
    print("  ✓ PayPal transactions (categorized)")
    print("  ✓ Storm Expenses")
    print("  ✓ All bank statement transactions (1,189 transactions)")
    
    print("\nWhat we NEED to ADD:")
    print("  1. Business expense categorization (FFJ Consulting LLC expenses)")
    print("     - Need to categorize transactions into: Software, Hardware, Office, Auto, etc.")
    print("  2. Rental property maintenance/improvement expenses")
    print("     - Need to identify which transactions are for which property")
    print("     - Categorize as: Maintenance Labor, Maintenance Materials, Improvement Labor, Improvement Materials")
    print("  3. Other income sources")
    print("     - Interest income")
    print("     - Dividend income")
    print("     - Capital gains")
    print("     - Social Security benefits")
    print("  4. Tax withholding details")
    print("     - Verify W-2 tax withholdings are captured correctly")
    print("  5. Any 1099 forms (1099-MISC, 1099-R, 1099-G, etc.)")
    print("     - Need to know if you received any 1099 forms")
    
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
