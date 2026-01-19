#!/usr/bin/env python3
"""
Analyze the Excel template structure to understand what needs to be filled.
"""

import sys
try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

def analyze_excel_template(file_path):
    """Analyze the Excel template structure."""
    print(f"Analyzing template: {file_path}")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        print(f"\nSheet names: {wb.sheetnames}")
        print("\n" + "=" * 60)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print("-" * 60)
            
            # Get dimensions
            print(f"Dimensions: {sheet.dimensions}")
            
            # Show first few rows with data
            print("\nFirst 20 rows:")
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                if any(cell is not None for cell in row):
                    print(f"Row {i}: {row}")
            
            # Look for headers
            print("\nLooking for headers (first non-empty row):")
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                if any(cell is not None for cell in row):
                    print(f"Row {i} (potential headers): {row}")
                    break
            
            # Check for merged cells
            if sheet.merged_cells:
                print(f"\nMerged cells: {list(sheet.merged_cells)}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    template_path = "/Users/fjabbari/@@@PUBLIC/@@@TAX2025/final001.xlsx"
    analyze_excel_template(template_path)
