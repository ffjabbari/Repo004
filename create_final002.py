#!/usr/bin/env python3
"""
Create final002.xlsx from database with all available information.
"""

import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path


def create_monthly_rental_income(db_path, tax_year=2025):
    """Create monthly rental income breakdown (annual / 12)."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Get rental income by property
    cursor.execute("""
        SELECT entity_name, total_amount
        FROM EntitySummary
        WHERE entity_name LIKE 'RENTAL-%' AND tax_year = ?
        ORDER BY entity_name
    """, (tax_year,))
    
    rental_data = {}
    for prop_entity, annual_amount in cursor.fetchall():
        prop_num = prop_entity.replace('RENTAL-', '')
        monthly = annual_amount / 12.0
        rental_data[prop_num] = {
            'annual': annual_amount,
            'monthly': monthly
        }
    
    conn.close()
    return rental_data


def identify_business_expenses(db_path, tax_year=2025):
    """Identify potential business expenses for manual review."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Business expense keywords
    business_keywords = {
        'Software/MIS': ['intellij', 'jetbrains', 'zoom', 'software', 'subscription', 'adobe', 'microsoft', 'slack', 'github'],
        'Hardware': ['apple', 'macbook', 'computer', 'laptop', 'monitor', 'keyboard', 'mouse', 'hardware', 'dell', 'hp'],
        'Office': ['office', 'supplies', 'staples', 'office depot', 'paper', 'printer', 'ink', 'toner', 'furniture'],
        'Auto': ['gas', 'fuel', 'exxon', 'shell', 'chevron', 'auto', 'car', 'vehicle', 'repair', 'parking', 'toll'],
        'Bank Fee': ['fee', 'service charge', 'overdraft', 'atm fee', 'monthly maintenance'],
        'Insurance/Medical': ['insurance', 'medical', 'health', 'dental', 'premium', 'blue cross', 'aetna'],
        'Meals/Entertainment': ['restaurant', 'dining', 'food', 'meal', 'starbucks', 'cafe', 'bar', 'entertainment'],
        'Professional Fees': ['udemy', 'course', 'training', 'education', 'certification', 'coursera'],
        'Phone/Internet': ['phone', 'verizon', 'at&t', 't-mobile', 'internet', 'wifi', 'comcast', 'spectrum'],
        'Travel Lodging': ['hotel', 'lodging', 'airbnb', 'booking', 'marriott', 'hilton'],
        'Travel Transportation': ['airline', 'flight', 'delta', 'united', 'train', 'taxi', 'rental car', 'hertz'],
        'Utilities': ['electric', 'gas utility', 'water', 'sewer', 'trash', 'utility']
    }
    
    # Get all debit transactions
    cursor.execute("""
        SELECT entity_name, description, amount, transaction_date
        FROM EntityDetail
        WHERE tax_year = ? AND amount < 0
        ORDER BY transaction_date
    """, (tax_year,))
    
    business_expenses = {category: [] for category in business_keywords.keys()}
    business_expenses['Uncategorized'] = []
    
    for entity, description, amount, date in cursor.fetchall():
        desc_lower = (description or '').lower()
        entity_lower = (entity or '').lower()
        combined = f"{desc_lower} {entity_lower}"
        
        categorized = False
        for category, keywords in business_keywords.items():
            if any(keyword in combined for keyword in keywords):
                business_expenses[category].append({
                    'entity': entity,
                    'description': description,
                    'amount': abs(amount),
                    'date': date
                })
                categorized = True
                break
        
        if not categorized:
            business_expenses['Uncategorized'].append({
                'entity': entity,
                'description': description,
                'amount': abs(amount),
                'date': date
            })
    
    conn.close()
    return business_expenses


def identify_rental_expenses(db_path, tax_year=2025):
    """Identify rental property expenses and divide among 6 rental properties."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Rental property keywords
    rental_keywords = [
        'home depot', 'lowes', 'homedepot', 'contractor', 'plumber', 'electrician',
        'handyman', 'repair', 'maintenance', 'hvac', 'roof', 'paint', 'supplies',
        'hardware', 'lumber', 'tools', 'appliance', 'furnace', 'ac', 'heating',
        'cooling', 'windows', 'doors', 'flooring', 'cabinets', 'renovation',
        'remodel', 'construction', 'installation', 'improvement', 'upgrade'
    ]
    
    # Property addresses that might appear in descriptions
    property_addresses = {
        '1029': ['1029', 'almont'],
        '1035': ['1035', 'almont'],
        '1108': ['1108', 'almont'],
        '5015': ['5015', 'hi-view', 'hiview'],
        '5952': ['5952', 'pershing'],
        '5956': ['5956', 'pershing']
    }
    
    # Get all debit transactions
    cursor.execute("""
        SELECT entity_name, description, amount, transaction_date
        FROM EntityDetail
        WHERE tax_year = ? AND amount < 0
        ORDER BY transaction_date
    """, (tax_year,))
    
    rental_expenses = {
        '1029': {'maintenance_labor': [], 'maintenance_materials': [], 'improvement_labor': [], 'improvement_materials': []},
        '1035': {'maintenance_labor': [], 'maintenance_materials': [], 'improvement_labor': [], 'improvement_materials': []},
        '1108': {'maintenance_labor': [], 'maintenance_materials': [], 'improvement_labor': [], 'improvement_materials': []},
        '5015': {'maintenance_labor': [], 'maintenance_materials': [], 'improvement_labor': [], 'improvement_materials': []},
        '5952': {'maintenance_labor': [], 'maintenance_materials': [], 'improvement_labor': [], 'improvement_materials': []},
        '5956': {'maintenance_labor': [], 'maintenance_materials': [], 'improvement_labor': [], 'improvement_materials': []},
        'unassigned': []
    }
    
    labor_keywords = ['labor', 'contractor', 'plumber', 'electrician', 'handyman', 'work', 'service']
    material_keywords = ['supplies', 'material', 'hardware', 'lumber', 'tools', 'parts', 'paint', 'appliance']
    improvement_keywords = ['renovation', 'remodel', 'construction', 'improvement', 'upgrade', 'hvac', 'roof', 'windows', 'doors', 'flooring', 'cabinets']
    
    for entity, description, amount, date in cursor.fetchall():
        desc_lower = (description or '').lower()
        entity_lower = (entity or '').lower()
        combined = f"{desc_lower} {entity_lower}"
        
        # Check if it's a rental-related expense
        is_rental = any(keyword in combined for keyword in rental_keywords)
        
        if is_rental:
            # Try to identify which property
            assigned_property = None
            for prop_num, addresses in property_addresses.items():
                if any(addr in combined for addr in addresses):
                    assigned_property = prop_num
                    break
            
            # If not identified, divide equally among 6 properties
            if not assigned_property:
                # Divide amount by 6
                divided_amount = abs(amount) / 6.0
                for prop_num in property_addresses.keys():
                    # Determine category
                    is_labor = any(kw in combined for kw in labor_keywords)
                    is_material = any(kw in combined for kw in material_keywords)
                    is_improvement = any(kw in combined for kw in improvement_keywords)
                    
                    if is_improvement:
                        if is_labor:
                            category = 'improvement_labor'
                        else:
                            category = 'improvement_materials'
                    else:
                        if is_labor:
                            category = 'maintenance_labor'
                        else:
                            category = 'maintenance_materials'
                    
                    rental_expenses[prop_num][category].append({
                        'entity': entity,
                        'description': description,
                        'amount': divided_amount,
                        'date': date
                    })
            else:
                # Assign to specific property
                is_labor = any(kw in combined for kw in labor_keywords)
                is_material = any(kw in combined for kw in material_keywords)
                is_improvement = any(kw in combined for kw in improvement_keywords)
                
                if is_improvement:
                    if is_labor:
                        category = 'improvement_labor'
                    else:
                        category = 'improvement_materials'
                else:
                    if is_labor:
                        category = 'maintenance_labor'
                    else:
                        category = 'maintenance_materials'
                
                rental_expenses[assigned_property][category].append({
                    'entity': entity,
                    'description': description,
                    'amount': abs(amount),
                    'date': date
                })
    
    conn.close()
    return rental_expenses


def create_final002():
    """Create final002.xlsx from template and database."""
    
    template_path = "/Users/fjabbari/@@@PUBLIC/@@@TAX2025/final001.xlsx"
    output_path = "/Users/fjabbari/@@@PUBLIC/@@@TAX2025/final002.xlsx"
    db_path = "tax_data_2025.db"
    tax_year = 2025
    
    print("=" * 80)
    print("CREATING final002.xlsx")
    print("=" * 80)
    
    # Load template
    print("\n1. Loading template...")
    wb = openpyxl.load_workbook(template_path)
    print(f"   ✓ Loaded template with {len(wb.sheetnames)} sheets")
    
    # Get data from database
    print("\n2. Gathering data from database...")
    
    # Monthly rental income
    rental_data = create_monthly_rental_income(db_path, tax_year)
    print(f"   ✓ Monthly rental income calculated for {len(rental_data)} properties")
    
    # Business expenses
    business_expenses = identify_business_expenses(db_path, tax_year)
    total_business = sum(sum(item['amount'] for item in items) for items in business_expenses.values())
    print(f"   ✓ Identified business expenses: ${total_business:,.2f} (for manual review)")
    
    # Rental expenses
    rental_expenses = identify_rental_expenses(db_path, tax_year)
    total_rental = sum(
        sum(item['amount'] for item in rental_expenses[prop][cat])
        for prop in ['1029', '1035', '1108', '5015', '5952', '5956']
        for cat in ['maintenance_labor', 'maintenance_materials', 'improvement_labor', 'improvement_materials']
    )
    print(f"   ✓ Identified rental expenses: ${total_rental:,.2f}")
    
    # Get W-2 and other data
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    cursor.execute("SELECT total_amount FROM EntitySummary WHERE entity_name = 'W-2 INCOME' AND tax_year = ?", (tax_year,))
    w2_result = cursor.fetchone()
    w2_wages = w2_result[0] if w2_result else 0
    
    cursor.execute("SELECT total_amount FROM EntitySummary WHERE entity_name = 'SSA TREAS' AND tax_year = ?", (tax_year,))
    ssa_result = cursor.fetchone()
    ssa_amount = ssa_result[0] if ssa_result else 0
    
    conn.close()
    
    # Fill template sheets
    print("\n3. Filling template sheets...")
    
    # Fill "1099misc+1099R+w2+1099G1 (2)" sheet
    if '1099misc+1099R+w2+1099G1 (2)' in wb.sheetnames:
        sheet = wb['1099misc+1099R+w2+1099G1 (2)']
        print("   Filling W-2 sheet...")
        
        # Find row to add data (usually row 2 or first empty row after headers)
        data_row = 2
        while sheet.cell(data_row, 1).value and data_row < 20:
            data_row += 1
        
        # W-2 data
        sheet.cell(data_row, 1).value = "W-2"  # DOC TYPE
        sheet.cell(data_row, 2).value = w2_wages  # W2 wages
        sheet.cell(data_row, 3).value = w2_wages  # ssn wages (same as W-2 for now)
        # Tax withholdings would go here - we have them but need to verify placement
        
        print(f"     ✓ Added W-2 wages: ${w2_wages:,.2f}")
    
    # Fill "worksheet-RENT INCOME" sheet
    if 'worksheet-RENT INCOME' in wb.sheetnames:
        sheet = wb['worksheet-RENT INCOME']
        print("   Filling rental income sheet...")
        
        # Find property columns and month rows
        # Properties are typically in row 2-4, months start around row 5
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        # Find month rows
        month_rows = {}
        for row_idx in range(5, 20):
            cell = sheet.cell(row_idx, 2)  # Column B typically has months
            if cell.value and str(cell.value).strip() in months:
                month_rows[str(cell.value).strip()] = row_idx
        
        # Find property columns (look for property numbers in row 2-4)
        prop_columns = {}
        for col_idx in range(3, 20):
            for row_idx in range(2, 5):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value:
                    cell_val = str(cell.value).strip()
                    for prop_num in rental_data.keys():
                        if prop_num in cell_val or cell_val == prop_num:
                            prop_columns[prop_num] = col_idx
                            break
        
        # Fill monthly rental income
        for prop_num, data in rental_data.items():
            if prop_num in prop_columns:
                col = prop_columns[prop_num]
                for month, row in month_rows.items():
                    sheet.cell(row, col).value = data['monthly']
        
        print(f"     ✓ Added monthly rental income for {len(rental_data)} properties")
    
    # Fill "FFJ Consulting LLC EXP" sheet
    if 'FFJ Consulting LLC EXP' in wb.sheetnames:
        sheet = wb['FFJ Consulting LLC EXP']
        print("   Filling business expenses sheet (for manual review)...")
        
        # Map categories to columns
        category_map = {
            'Software/MIS': 'Software\n and MIS(INTELLIJ, ZOOM, OTHER...)',
            'Hardware': 'Hardware Computer\npurchased',
            'Office': 'office exp',
            'Auto': 'Auto EXP',
            'Bank Fee': 'Bank Fee',
            'Insurance/Medical': 'Insurnce\nliab\u2028Medical',
            'Meals/Entertainment': 'Meal\u2028\nEntertainment',
            'Professional Fees': 'Professional\u2028\nFee ( UDEMY Courses)',
            'Phone/Internet': 'Phone and\nComm\u2028+ internet\nExp',
            'Travel Lodging': 'Travel\u2028Exp(Lodging)',
            'Travel Transportation': 'Travel \nTransportation',
            'Utilities': 'Utilities'
        }
        
        # Find column for each category
        col_map = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).strip()
                for cat, header_text in category_map.items():
                    if header_text in header or header in header_text:
                        col_map[cat] = col_idx
        
        # Fill data row (usually row 2)
        data_row = 2
        for category, items in business_expenses.items():
            if category in col_map and category != 'Uncategorized':
                total = sum(item['amount'] for item in items)
                col = col_map[category]
                sheet.cell(data_row, col).value = total
        
        print(f"     ✓ Added business expenses (marked for manual review)")
    
    # Fill "worksheet Rental DetailRecieptA" sheet
    if 'worksheet Rental DetailRecieptA' in wb.sheetnames:
        sheet = wb['worksheet Rental DetailRecieptA']
        print("   Filling rental expenses sheet...")
        
        # Find category columns
        category_cols = {
            'maintenance_labor': 'mamaint/lab\nALL LABORS GO HERE',
            'maintenance_materials': 'main/mat',
            'improvement_labor': 'imp/lab',
            'improvement_materials': 'imp/mat'
        }
        
        col_map = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header = str(cell.value).strip()
                for cat, header_text in category_cols.items():
                    if header_text.lower() in header.lower() or header.lower() in header_text.lower():
                        col_map[cat] = col_idx
        
        # Find property rows (properties in column A)
        prop_rows = {}
        for row_idx in range(2, 50):
            cell = sheet.cell(row_idx, 1)  # Column A
            if cell.value:
                cell_val = str(cell.value).strip()
                for prop_num in ['1029', '1035', '1108', '5015', '5952', '5956']:
                    if prop_num in cell_val or cell_val.endswith(prop_num):
                        prop_rows[prop_num] = row_idx
                        break
        
        # Fill rental expenses
        for prop_num in ['1029', '1035', '1108', '5015', '5952', '5956']:
            if prop_num in prop_rows:
                row = prop_rows[prop_num]
                for cat, col in col_map.items():
                    total = sum(item['amount'] for item in rental_expenses[prop_num][cat])
                    if total > 0:
                        sheet.cell(row, col).value = total
        
        print(f"     ✓ Added rental expenses for 6 properties")
    
    # Save file
    print(f"\n4. Saving final002.xlsx...")
    wb.save(output_path)
    wb.close()
    
    print(f"   ✓ Saved to: {output_path}")
    
    print("\n" + "=" * 80)
    print("final002.xlsx CREATED SUCCESSFULLY!")
    print("=" * 80)
    
    print("\nSummary:")
    print(f"  - W-2 Income: ${w2_wages:,.2f}")
    print(f"  - Rental Income: {len(rental_data)} properties, monthly breakdown added")
    print(f"  - Business Expenses: ${total_business:,.2f} (for manual review)")
    print(f"  - Rental Expenses: ${total_rental:,.2f} (divided among 6 properties)")
    print(f"  - Social Security: ${ssa_amount:,.2f}")
    
    print("\nNote: Business expenses are categorized but need manual review.")
    print("      Rental expenses are divided equally among 6 properties where not specifically identified.")


if __name__ == "__main__":
    create_final002()
