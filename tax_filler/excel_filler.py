"""
Fill the Excel tax template with categorized transactions.
"""

import openpyxl
from typing import List, Dict, Optional
from datetime import datetime
from tax_filler.categorizer import CategorizedTransaction
from tax_filler.template_analyzer import TemplateStructure


class ExcelFiller:
    """Fill Excel template with categorized transactions."""
    
    def __init__(self, template_path: str, structure: TemplateStructure):
        """
        Initialize Excel filler.
        
        Args:
            template_path: Path to the Excel template
            structure: TemplateStructure object
        """
        self.template_path = template_path
        self.structure = structure
        self.wb = None
    
    def fill(self, categorized_transactions: List[CategorizedTransaction], output_path: Optional[str] = None):
        """
        Fill the template with categorized transactions.
        
        Args:
            categorized_transactions: List of categorized transactions
            output_path: Optional output path (defaults to template_path with _filled suffix)
        """
        # Load workbook
        self.wb = openpyxl.load_workbook(self.template_path)
        
        # Fill different sections
        self._fill_consulting_expenses(categorized_transactions)
        self._fill_rental_income(categorized_transactions)
        self._fill_rental_expenses(categorized_transactions)
        
        # Save
        if output_path is None:
            output_path = self.template_path.replace('.xlsx', '_filled.xlsx')
        
        self.wb.save(output_path)
        print(f"Filled template saved to: {output_path}")
    
    def _fill_consulting_expenses(self, transactions: List[CategorizedTransaction]):
        """Fill consulting expenses sheet."""
        if self.structure.consulting_expenses_sheet not in self.wb.sheetnames:
            print(f"Sheet '{self.structure.consulting_expenses_sheet}' not found")
            return
        
        sheet = self.wb[self.structure.consulting_expenses_sheet]
        
        # Get column mappings
        column_map = self._get_consulting_column_map(sheet)
        if not column_map:
            print("Could not find consulting expense columns")
            return
        
        # Filter business expenses
        business_expenses = [t for t in transactions if t.is_business_expense]
        
        # Group by category
        expenses_by_category = {}
        for trans in business_expenses:
            category = trans.category
            if category not in expenses_by_category:
                expenses_by_category[category] = []
            expenses_by_category[category].append(trans)
        
        # Find the data row (usually row 2)
        data_row = 2
        
        # Fill amounts by category
        for category, trans_list in expenses_by_category.items():
            if category in column_map:
                col_letter = column_map[category]
                total = sum(abs(t.transaction.amount) for t in trans_list)
                
                # Find existing value or set new
                cell = sheet[f"{col_letter}{data_row}"]
                current_value = cell.value or 0
                if isinstance(current_value, (int, float)):
                    cell.value = current_value + total
                else:
                    cell.value = total
        
        # Update total row if exists
        self._update_total_row(sheet, column_map, data_row)
    
    def _fill_rental_income(self, transactions: List[CategorizedTransaction]):
        """Fill rental income sheet."""
        if self.structure.rental_income_sheet not in self.wb.sheetnames:
            print(f"Sheet '{self.structure.rental_income_sheet}' not found")
            return
        
        sheet = self.wb[self.structure.rental_income_sheet]
        
        # Filter rental income
        rental_income = [t for t in transactions if t.is_rental_income]
        
        # Group by property and month
        income_by_property_month = {}
        for trans in rental_income:
            prop = trans.property_address or 'UNKNOWN'
            month = trans.transaction.date.month
            
            key = (prop, month)
            if key not in income_by_property_month:
                income_by_property_month[key] = []
            income_by_property_month[key].append(trans)
        
        # Find property columns and month rows
        prop_col_map = self._get_property_column_map(sheet)
        month_row_map = self._get_month_row_map(sheet)
        
        # Fill amounts
        for (prop, month), trans_list in income_by_property_month.items():
            if prop in prop_col_map and month in month_row_map:
                col_letter = prop_col_map[prop]
                row_num = month_row_map[month]
                
                total = sum(t.transaction.amount for t in trans_list)
                cell = sheet[f"{col_letter}{row_num}"]
                current_value = cell.value or 0
                if isinstance(current_value, (int, float)):
                    cell.value = current_value + total
                else:
                    cell.value = total
    
    def _fill_rental_expenses(self, transactions: List[CategorizedTransaction]):
        """Fill rental expenses sheet."""
        if self.structure.rental_expenses_sheet not in self.wb.sheetnames:
            print(f"Sheet '{self.structure.rental_expenses_sheet}' not found")
            return
        
        sheet = self.wb[self.structure.rental_expenses_sheet]
        
        # Filter rental expenses
        rental_expenses = [t for t in transactions if t.is_rental_expense]
        
        # Group by property and category
        expenses_by_property_category = {}
        for trans in rental_expenses:
            prop = trans.property_address or 'UNKNOWN'
            category = trans.category
            
            key = (prop, category)
            if key not in expenses_by_property_category:
                expenses_by_property_category[key] = []
            expenses_by_property_category[key].append(trans)
        
        # Find property rows and category columns
        prop_row_map = self._get_property_row_map(sheet)
        category_col_map = self._get_category_column_map(sheet)
        
        # Fill amounts
        for (prop, category), trans_list in expenses_by_property_category.items():
            if prop in prop_row_map and category in category_col_map:
                row_num = prop_row_map[prop]
                col_letter = category_col_map[category]
                
                total = sum(abs(t.transaction.amount) for t in trans_list)
                cell = sheet[f"{col_letter}{row_num}"]
                current_value = cell.value or 0
                if isinstance(current_value, (int, float)):
                    cell.value = current_value + total
                else:
                    cell.value = total
    
    def _get_consulting_column_map(self, sheet) -> Dict[str, str]:
        """Get column letter mapping for consulting expense categories."""
        column_map = {}
        
        # Check header row (row 1)
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header_text = str(cell.value).strip()
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                
                # Match against known categories
                for category in self.structure.consulting_expense_columns.keys():
                    if category.lower() in header_text.lower() or header_text.lower() in category.lower():
                        column_map[category] = col_letter
        
        return column_map
    
    def _get_property_column_map(self, sheet) -> Dict[str, str]:
        """Get column letter mapping for rental properties."""
        column_map = {}
        
        # Properties are typically in row 2-4, columns C onwards
        for row_idx in range(2, 5):
            row = sheet[row_idx]
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    prop_str = str(cell.value).strip()
                    if prop_str in self.structure.rental_properties:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        column_map[prop_str] = col_letter
        
        return column_map
    
    def _get_month_row_map(self, sheet) -> Dict[int, int]:
        """Get row number mapping for months."""
        month_map = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        
        row_map = {}
        
        # Months are typically in column B, starting from row 5
        for row_idx in range(5, 20):
            cell = sheet[f"B{row_idx}"]
            if cell.value:
                month_str = str(cell.value).strip().lower()
                if month_str in month_map:
                    row_map[month_map[month_str]] = row_idx
        
        return row_map
    
    def _get_property_row_map(self, sheet) -> Dict[str, int]:
        """Get row number mapping for rental properties in expenses sheet."""
        row_map = {}
        
        # Properties are in column A
        for row_idx in range(2, 200):
            cell = sheet[f"A{row_idx}"]
            if cell.value:
                prop_str = str(cell.value).strip()
                if prop_str in self.structure.rental_properties or 'TOTAL' in prop_str.upper():
                    row_map[prop_str] = row_idx
        
        return row_map
    
    def _get_category_column_map(self, sheet) -> Dict[str, str]:
        """Get column letter mapping for expense categories."""
        column_map = {}
        
        # Categories are in row 1
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                category_text = str(cell.value).strip().lower()
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                
                # Map common category names
                if 'maint' in category_text and 'lab' in category_text:
                    column_map['mamaint/lab'] = col_letter
                elif 'main' in category_text and 'mat' in category_text:
                    column_map['main/mat'] = col_letter
                elif 'imp' in category_text and 'lab' in category_text:
                    column_map['imp/lab'] = col_letter
                elif 'imp' in category_text and 'mat' in category_text:
                    column_map['imp/mat'] = col_letter
        
        return column_map
    
    def _update_total_row(self, sheet, column_map: Dict[str, str], data_row: int):
        """Update total row if it exists."""
        # Look for TOTAL row (usually after data rows)
        total_row = None
        for row_idx in range(data_row, data_row + 10):
            cell = sheet[f"A{row_idx}"]
            if cell.value and 'TOTAL' in str(cell.value).upper():
                total_row = row_idx
                break
        
        if total_row:
            # Calculate sum for each category column
            for category, col_letter in column_map.items():
                cell = sheet[f"{col_letter}{total_row}"]
                # Set formula or value
                cell.value = f"=SUM({col_letter}{data_row}:{col_letter}{total_row-1})"
    
    def close(self):
        """Close the workbook."""
        if self.wb:
            self.wb.close()
