#!/usr/bin/env python3
"""
Detailed analysis of template requirements.
"""

import openpyxl
import sqlite3
from collections import defaultdict


def detailed_template_analysis():
    """Detailed analysis of what the template needs."""
    
    template_path = "/Users/fjabbari/@@@PUBLIC/@@@TAX2025/final001.xlsx"
    db_path = "tax_data_2025.db"
    
    wb = openpyxl.load_workbook(template_path, data_only=True)
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    print("=" * 80)
    print("DETAILED TEMPLATE REQUIREMENTS ANALYSIS")
    print("=" * 80)
    
    # 1. Analyze "1099misc+1099R+w2+1099G1 (2)" sheet
    print("\n1. INCOME & TAX WITHHOLDING SHEET:")
    print("-" * 80)
    if '1099misc+1099R+w2+1099G1 (2)' in wb.sheetnames:
        sheet = wb['1099misc+1099R+w2+1099G1 (2)']
        print("   Sheet: 1099misc+1099R+w2+1099G1 (2)")
        
        # Get headers
        headers = [cell.value for cell in sheet[1] if cell.value]
        print(f"   Columns: {headers[:10]}")
        
        # Check what we have
        cursor.execute("SELECT total_amount FROM EntitySummary WHERE entity_name = 'W-2 INCOME'")
        w2_result = cursor.fetchone()
        if w2_result:
            print(f"   ✓ W-2 Wages in database: ${w2_result[0]:,.2f}")
        else:
            print("   ✗ W-2 Wages: MISSING")
        
        # Check for other income types
        print("   ? Need: Interest income, Dividends, 1099-MISC, 1099-R, 1099-G")
    
    # 2. Analyze "FFJ Consulting LLC EXP" sheet
    print("\n2. BUSINESS EXPENSES SHEET (FFJ Consulting LLC):")
    print("-" * 80)
    if 'FFJ Consulting LLC EXP' in wb.sheetnames:
        sheet = wb['FFJ Consulting LLC EXP']
        print("   Sheet: FFJ Consulting LLC EXP")
        
        # Get expense categories from row 1
        categories = []
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value and len(str(cell.value).strip()) > 3:
                cat = str(cell.value).strip()
                categories.append(cat)
        
        print(f"   Expense categories needed:")
        for cat in categories[:15]:
            if cat and 'Item description' not in cat and 'TOTAL' not in cat:
                print(f"     - {cat}")
        
        # Check if we have business expenses categorized
        cursor.execute("""
            SELECT COUNT(*) FROM EntitySummary
            WHERE entity_name LIKE 'FFJ-%' OR category = 'Business Expense'
        """)
        has_business = cursor.fetchone()[0] > 0
        if has_business:
            print(f"   ✓ Some business expenses found")
        else:
            print(f"   ✗ Business expenses NOT categorized yet")
            print(f"     Need to categorize transactions into the above categories")
    
    # 3. Analyze "worksheet-RENT INCOME" sheet
    print("\n3. RENTAL INCOME SHEET:")
    print("-" * 80)
    if 'worksheet-RENT INCOME' in wb.sheetnames:
        sheet = wb['worksheet-RENT INCOME']
        print("   Sheet: worksheet-RENT INCOME")
        
        # Check rental income by property
        cursor.execute("""
            SELECT entity_name, total_amount
            FROM EntitySummary
            WHERE entity_name LIKE 'RENTAL-%'
            ORDER BY entity_name
        """)
        rental_properties = cursor.fetchall()
        
        print(f"   Rental properties in database: {len(rental_properties)}")
        for prop, amount in rental_properties:
            print(f"     ✓ {prop}: ${amount:,.2f}")
        
        # Check if template has monthly breakdown
        print("   ? Need: Monthly rental income breakdown (Jan-Dec)")
    
    # 4. Analyze "worksheet Rental DetailRecieptA" sheet
    print("\n4. RENTAL EXPENSES SHEET:")
    print("-" * 80)
    if 'worksheet Rental DetailRecieptA' in wb.sheetnames:
        sheet = wb['worksheet Rental DetailRecieptA']
        print("   Sheet: worksheet Rental DetailRecieptA")
        
        # Get categories
        categories = []
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                cat = str(cell.value).strip()
                if cat and len(cat) > 2:
                    categories.append(cat)
        
        print(f"   Rental expense categories needed:")
        for cat in categories[:10]:
            if 'desc' not in cat.lower() and 'TOTAL' not in cat:
                print(f"     - {cat}")
        
        # Check if we have rental expenses
        cursor.execute("""
            SELECT COUNT(*) FROM EntitySummary
            WHERE entity_name LIKE 'PROPERTY-%' AND entity_name LIKE '%-TAX'
        """)
        has_rental_expenses = cursor.fetchone()[0] > 0
        if has_rental_expenses:
            print(f"   ✓ Property taxes found")
        else:
            print(f"   ✗ Rental property expenses NOT categorized")
        
        print(f"   ? Need: Maintenance/Improvement expenses by property")
        print(f"     - Maintenance Labor")
        print(f"     - Maintenance Materials")
        print(f"     - Improvement Labor")
        print(f"     - Improvement Materials")
    
    # 5. Check for other income sources
    print("\n5. OTHER INCOME SOURCES:")
    print("-" * 80)
    
    # Check database for interest, dividends, etc.
    cursor.execute("""
        SELECT entity_name, total_amount
        FROM EntitySummary
        WHERE entity_name LIKE '%INTEREST%' 
           OR entity_name LIKE '%DIVIDEND%'
           OR entity_name LIKE '%SSA%'
           OR entity_name LIKE '%SOCIAL SECURITY%'
        ORDER BY ABS(total_amount) DESC
    """)
    other_income = cursor.fetchall()
    
    if other_income:
        print("   Found in database:")
        for entity, amount in other_income:
            print(f"     ✓ {entity}: ${amount:,.2f}")
    else:
        print("   ✗ No other income sources found")
        print("   ? Need: Interest income, Dividends, Social Security benefits")
    
    # 6. Tax withholdings
    print("\n6. TAX WITHHOLDINGS:")
    print("-" * 80)
    print("   ? Need to verify W-2 tax withholdings are captured")
    print("     - Federal tax withheld")
    print("     - State tax withheld")
    print("     - Social Security tax withheld")
    print("     - Medicare tax withheld")
    
    wb.close()
    conn.close()
    
    # Final summary
    print("\n" + "=" * 80)
    print("WHAT WE NEED FROM YOU:")
    print("=" * 80)
    
    print("\n1. BUSINESS EXPENSE CATEGORIZATION:")
    print("   - Review transactions and tell me which ones are business expenses")
    print("   - Or provide rules/keywords to auto-categorize")
    print("   - Categories needed: Software, Hardware, Office, Auto, Bank Fees,")
    print("     Insurance, Meals, Professional Fees, Phone/Internet, Travel, Utilities")
    
    print("\n2. RENTAL PROPERTY EXPENSES:")
    print("   - Which transactions are for property maintenance/improvements?")
    print("   - Which property does each expense belong to?")
    print("   - Categorize as: Maintenance Labor, Maintenance Materials,")
    print("     Improvement Labor, Improvement Materials")
    
    print("\n3. OTHER INCOME:")
    print("   - Interest income (from bank statements)")
    print("   - Dividend income")
    print("   - Capital gains")
    print("   - Social Security benefits (if not already captured)")
    print("   - Any 1099 forms received (1099-MISC, 1099-R, 1099-G, etc.)")
    
    print("\n4. VERIFICATION:")
    print("   - Verify W-2 tax withholdings are correct")
    print("   - Verify rental income amounts are correct")
    print("   - Any other income or expenses not in bank statements?")
    
    print("\n" + "=" * 80)


if __name__ == "__main__":
    detailed_template_analysis()
