#!/usr/bin/env python3
"""
Create final1099.xlsx with PayPal payment summaries for 1099 preparation.
"""

import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_final1099():
    """Create final1099.xlsx with PayPal payment summaries."""
    
    db_path = "tax_data_2025.db"
    output_path = "/Users/fjabbari/@@@PUBLIC/@@@TAX2025/final1099.xlsx"
    tax_year = 2025
    
    print("=" * 80)
    print("CREATING final1099.xlsx")
    print("=" * 80)
    
    # Connect to database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Get all PayPal entities
    print("\n1. Gathering PayPal payment data...")
    cursor.execute("""
        SELECT 
            entity_name,
            total_amount,
            transaction_count,
            first_transaction_date,
            last_transaction_date
        FROM EntitySummary
        WHERE entity_name LIKE 'PAYPAL-%' AND tax_year = ?
        ORDER BY ABS(total_amount) DESC
    """, (tax_year,))
    
    paypal_summaries = cursor.fetchall()
    print(f"   ✓ Found {len(paypal_summaries)} PayPal entities")
    
    # Get detail transactions for each PayPal entity
    print("\n2. Gathering transaction details...")
    paypal_details = {}
    for entity_name, total, count, first_date, last_date in paypal_summaries:
        cursor.execute("""
            SELECT transaction_date, description, amount
            FROM EntityDetail
            WHERE entity_name = ? AND tax_year = ?
            ORDER BY transaction_date
        """, (entity_name, tax_year))
        
        details = cursor.fetchall()
        paypal_details[entity_name] = details
        print(f"   ✓ {entity_name}: {len(details)} transactions")
    
    conn.close()
    
    # Create Excel workbook
    print("\n3. Creating Excel file...")
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet
    summary_sheet = wb.create_sheet("PayPal Summary", 0)
    
    # Summary headers
    headers = [
        'PayPal Entity',
        'Total Amount',
        'Transaction Count',
        'First Payment Date',
        'Last Payment Date',
        'Needs 1099?',
        'Notes'
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = summary_sheet.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write summary data
    total_all = 0.0
    total_count = 0
    
    for row_idx, (entity_name, total, count, first_date, last_date) in enumerate(paypal_summaries, 2):
        # Entity name
        summary_sheet.cell(row_idx, 1).value = entity_name
        
        # Total amount
        summary_sheet.cell(row_idx, 2).value = abs(total)
        summary_sheet.cell(row_idx, 2).number_format = '$#,##0.00'
        
        # Transaction count
        summary_sheet.cell(row_idx, 3).value = count
        
        # Dates
        summary_sheet.cell(row_idx, 4).value = first_date
        summary_sheet.cell(row_idx, 5).value = last_date
        
        # Needs 1099? (empty for user to fill)
        summary_sheet.cell(row_idx, 6).value = ""
        
        # Notes (empty for user to fill)
        summary_sheet.cell(row_idx, 7).value = ""
        
        total_all += abs(total)
        total_count += count
    
    # Add total row
    total_row = len(paypal_summaries) + 2
    summary_sheet.cell(total_row, 1).value = "TOTAL"
    summary_sheet.cell(total_row, 1).font = Font(bold=True)
    summary_sheet.cell(total_row, 2).value = total_all
    summary_sheet.cell(total_row, 2).number_format = '$#,##0.00'
    summary_sheet.cell(total_row, 2).font = Font(bold=True)
    summary_sheet.cell(total_row, 3).value = total_count
    summary_sheet.cell(total_row, 3).font = Font(bold=True)
    
    # Adjust column widths
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 18
    summary_sheet.column_dimensions['D'].width = 18
    summary_sheet.column_dimensions['E'].width = 18
    summary_sheet.column_dimensions['F'].width = 15
    summary_sheet.column_dimensions['G'].width = 40
    
    # Create detail sheet for each PayPal entity
    print("\n4. Creating detail sheets...")
    
    for entity_name, total, count, first_date, last_date in paypal_summaries:
        # Create sheet name (Excel sheet names have limitations)
        sheet_name = entity_name.replace('PAYPAL-', '')[:31]  # Excel limit is 31 chars
        
        detail_sheet = wb.create_sheet(sheet_name)
        
        # Write entity header
        detail_sheet.cell(1, 1).value = f"PayPal Entity: {entity_name}"
        detail_sheet.cell(1, 1).font = Font(bold=True, size=14)
        detail_sheet.merge_cells('A1:D1')
        
        detail_sheet.cell(2, 1).value = f"Total Amount: ${abs(total):,.2f}"
        detail_sheet.cell(2, 1).font = Font(bold=True)
        detail_sheet.cell(2, 2).value = f"Transactions: {count}"
        detail_sheet.cell(2, 2).font = Font(bold=True)
        detail_sheet.cell(2, 3).value = f"Date Range: {first_date} to {last_date}"
        detail_sheet.cell(2, 3).font = Font(bold=True)
        
        # Detail headers
        detail_headers = ['Date', 'Description', 'Amount', 'Notes']
        for col_idx, header in enumerate(detail_headers, 1):
            cell = detail_sheet.cell(4, col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write transaction details
        details = paypal_details[entity_name]
        for row_idx, (trans_date, description, amount) in enumerate(details, 5):
            detail_sheet.cell(row_idx, 1).value = trans_date
            detail_sheet.cell(row_idx, 2).value = description
            detail_sheet.cell(row_idx, 3).value = abs(amount)
            detail_sheet.cell(row_idx, 3).number_format = '$#,##0.00'
            detail_sheet.cell(row_idx, 4).value = ""  # Notes column for user
        
        # Add total row
        total_detail_row = len(details) + 5
        detail_sheet.cell(total_detail_row, 2).value = "TOTAL"
        detail_sheet.cell(total_detail_row, 2).font = Font(bold=True)
        detail_sheet.cell(total_detail_row, 3).value = abs(total)
        detail_sheet.cell(total_detail_row, 3).number_format = '$#,##0.00'
        detail_sheet.cell(total_detail_row, 3).font = Font(bold=True)
        
        # Adjust column widths
        detail_sheet.column_dimensions['A'].width = 12
        detail_sheet.column_dimensions['B'].width = 80
        detail_sheet.column_dimensions['C'].width = 15
        detail_sheet.column_dimensions['D'].width = 30
        
        print(f"   ✓ Created detail sheet for {entity_name}")
    
    # Save file
    print(f"\n5. Saving final1099.xlsx...")
    wb.save(output_path)
    wb.close()
    
    print(f"   ✓ Saved to: {output_path}")
    
    print("\n" + "=" * 80)
    print("final1099.xlsx CREATED SUCCESSFULLY!")
    print("=" * 80)
    
    print(f"\nSummary:")
    print(f"  - Total PayPal entities: {len(paypal_summaries)}")
    print(f"  - Total transactions: {total_count}")
    print(f"  - Total amount: ${total_all:,.2f}")
    print(f"\nFile contains:")
    print(f"  - Summary sheet with all PayPal entities (one line each)")
    print(f"  - Detail sheet for each PayPal entity with all transactions")
    print(f"  - 'Needs 1099?' column for you to mark which ones need 1099 forms")
    print(f"  - Notes column for your comments")


if __name__ == "__main__":
    create_final1099()
